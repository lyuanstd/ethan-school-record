from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _join_list(values):
    """
    Convert a list into a comma-separated string.
    """
    if not values:
        return ""

    return ", ".join(
        str(value).strip()
        for value in values
        if str(value).strip()
    )


def _format_additional_items(items):
    """
    Convert additional_items into one readable Excel cell.

    Example:
        Milk: Drank extra milk | Temperature: Normal
    """
    if not items:
        return ""

    formatted_items = []

    for item in items:
        field = str(
            item.get("field", "")
        ).strip()

        value = str(
            item.get("value", "")
        ).strip()

        if field and value:
            formatted_items.append(
                f"{field}: {value}"
            )

        elif field:
            formatted_items.append(field)

        elif value:
            formatted_items.append(value)

    return " | ".join(formatted_items)


def create_excel(records):
    """
    Create an Excel workbook from school records.

    Structured Daily Routine fields stored inside
    routine_data are expanded into separate columns.

    Returns:
        BytesIO object containing the Excel file.
    """

    rows = []

    for record in records:

        routine_data = (
            record.get("routine_data")
            or {}
        )

        rows.append(
            {
                "Date": (
                    record.get(
                        "record_date",
                        "",
                    )
                ),

                "English Course": (
                        record.get(
                            "english_course",
                            "",
                        )
                        or ""
                ),

                "Chinese Course": (
                        record.get(
                            "chinese_course",
                            "",
                        )
                        or ""
                ),

                "Daily Summary": (
                        record.get(
                            "daily_summary",
                            "",
                        )
                        or ""
                ),

                "Mood": _join_list(
                    routine_data.get(
                        "mood",
                        [],
                    )
                ),

                "Favourite Activities": (
                    _join_list(
                        routine_data.get(
                            "favorite_activities",
                            [],
                        )
                    )
                ),

                "Bowel Movement": (
                        routine_data.get(
                            "bowel_movement",
                            "",
                        )
                        or ""
                ),

                "Morning Snack": (
                        routine_data.get(
                            "morning_snack",
                            "",
                        )
                        or ""
                ),

                "Lunch": (
                        routine_data.get(
                            "lunch",
                            "",
                        )
                        or ""
                ),

                "Afternoon Snack": (
                        routine_data.get(
                            "afternoon_snack",
                            "",
                        )
                        or ""
                ),

                "Nap": (
                        routine_data.get(
                            "nap",
                            "",
                        )
                        or ""
                ),

                "Extra Diapers": (
                        routine_data.get(
                            "extra_diapers",
                            "",
                        )
                        or ""
                ),

                "Extra Clothes": (
                        routine_data.get(
                            "extra_clothes",
                            "",
                        )
                        or ""
                ),

                "Other": (
                        routine_data.get(
                            "other",
                            "",
                        )
                        or ""
                ),

                "Additional Items": (
                    _format_additional_items(
                        routine_data.get(
                            "additional_items",
                            [],
                        )
                    )
                ),

                "Daily Routine Image": (
                        record.get(
                            "routine_image",
                            "",
                        )
                        or ""
                ),
            }
        )

    dataframe = pd.DataFrame(rows)

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:

        dataframe.to_excel(
            writer,
            sheet_name="School Records",
            index=False,
        )

        worksheet = writer.book[
            "School Records"
        ]

        # ------------------------------------------
        # Header formatting
        # ------------------------------------------

        for cell in worksheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # ------------------------------------------
        # Body formatting
        # ------------------------------------------

        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
        ):

            for cell in row:

                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        # ------------------------------------------
        # Column widths
        # ------------------------------------------

        column_widths = {
            1: 14,  # Date
            2: 60,  # English Course
            3: 60,  # Chinese Course
            4: 50,  # Daily Summary
            5: 28,  # Mood
            6: 35,  # Favourite Activities
            7: 22,  # Bowel Movement
            8: 25,  # Morning Snack
            9: 25,  # Lunch
            10: 25,  # Afternoon Snack
            11: 20,  # Nap
            12: 20,  # Extra Diapers
            13: 20,  # Extra Clothes
            14: 35,  # Other
            15: 45,  # Additional Items
            16: 30,  # Daily Routine Image
        }

        for (
            column_number,
            width,
        ) in column_widths.items():

            column_letter = (
                get_column_letter(
                    column_number
                )
            )

            worksheet.column_dimensions[
                column_letter
            ].width = width

        # ------------------------------------------
        # Freeze header row
        # ------------------------------------------

        worksheet.freeze_panes = "A2"

        # ------------------------------------------
        # Optional autofilter
        # ------------------------------------------

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    output.seek(0)

    return output